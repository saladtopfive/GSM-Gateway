#!/usr/bin/env python3
import serial
import time
import datetime
import pytz
import os
import logging
import sys
import re
from openpyxl import load_workbook

PORT = "/dev/serial/by-id/usb-SimTech__Incorporated_SimTech__Incorporated_0123456789ABCDEF-if04-port0"
BAUD = 115200

SCHEDULE_CHECK_INTERVAL = 30
SMS_CHECK_INTERVAL = 1

LOCAL_TZ = pytz.timezone("Europe/Warsaw")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(SCRIPT_DIR, "schedule.xlsx")

PHONE_REGEX = re.compile(r"^(?:\+48)?\d{9}$")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    stream=sys.stdout,
)

log = logging.getLogger("gsm-forwarder")


def open_modem():
    try:
        ser = serial.Serial(PORT, BAUD, timeout=2)
        log.info("Polaczono z modemem")
        return ser
    except Exception as e:
        log.error("Nie mozna otworzyc portu: %s", e)
        return None


def send_at(ser, cmd, delay=0.3, log_cmd=False):
    ser.write((cmd + "\r").encode())
    time.sleep(delay)
    resp = ser.read_all().decode(errors="ignore").strip()
    if log_cmd:
        log.info("AT> %s | %s", cmd, resp)
    return resp


def enable_forwarding(ser, number):
    log.info("Ustawiam przekierowanie rozmow na %s", number)
    send_at(ser, f'AT+CCFC=0,3,"{number}",145', log_cmd=True)


def disable_forwarding(ser):
    log.info("Wylaczam przekierowanie rozmow")
    send_at(ser, "AT+CCFC=0,0", log_cmd=True)


def gsm_sanitize(text):
    replacements = {
        "ą": "a","ć": "c","ę": "e","ł": "l",
        "ń": "n","ó": "o","ś": "s",
        "ź": "z","ż": "z",
        "Ą": "A","Ć": "C","Ę": "E","Ł": "L",
        "Ń": "N","Ó": "O","Ś": "S",
        "Ź": "Z","Ż": "Z",
    }
    for pl, ascii_char in replacements.items():
        text = text.replace(pl, ascii_char)
    return text


def is_ucs2_hex(s):
    if len(s) % 4 != 0:
        return False
    return all(c in "0123456789ABCDEFabcdef" for c in s)


def decode_ucs2(hex_string):
    try:
        return bytes.fromhex(hex_string).decode("utf-16-be")
    except:
        return hex_string


def send_sms(ser, number, text):
    try:
        clean = gsm_sanitize(text)
        log.info("Wysylam SMS na %s", number)

        ser.write(f'AT+CMGS="{number}"\r'.encode())

        buffer = ""
        start = time.time()
        while ">" not in buffer:
            buffer += ser.read_all().decode(errors="ignore")
            if time.time() - start > 5:
                log.error("Timeout oczekiwania na >")
                return False
            time.sleep(0.1)

        ser.write(clean.encode("ascii", errors="ignore") + b"\x1A")

        response = ""
        start = time.time()
        while True:
            response += ser.read_all().decode(errors="ignore")

            if "+CMGS:" in response and "OK" in response:
                log.info("SMS wyslany poprawnie")
                return True

            if "ERROR" in response or "+CMS ERROR" in response:
                log.error("Blad wysylki SMS: %s", response.strip())
                return False

            if time.time() - start > 15:
                log.error("Timeout potwierdzenia SMS")
                return False

            time.sleep(0.2)

    except Exception as e:
        log.error("Wyjatek przy wysylaniu SMS: %s", e)
        return False


def process_all_sms(ser, forward_number):
    resp = send_at(ser, 'AT+CMGL="ALL"')

    if "+CMGL:" not in resp:
        return

    blocks = resp.split("+CMGL:")

    for block in blocks[1:]:
        lines = [l.strip() for l in block.splitlines() if l.strip()]
        if not lines:
            continue

        header = lines[0]
        body = lines[1] if len(lines) > 1 else ""

        parts = header.split(",")
        index = parts[0].strip()
        status = parts[1].replace('"', '').strip()

        sender_match = re.search(r'"(\+?\d+)"', header)
        if not sender_match:
            continue

        sender = sender_match.group(1)

        if is_ucs2_hex(body):
            body = decode_ucs2(body)

        log.info("SMS idx=%s status=%s od %s: %s", index, status, sender, body)

        if not forward_number:
            continue

        forward_text = f"{sender}:\n{body}"
        success = send_sms(ser, forward_number, forward_text)

        if success:
            send_at(ser, f"AT+CMGD={index}")
            log.info("SMS idx=%s usuniety po przekierowaniu", index)
        else:
            log.warning("Nie udalo sie przekierowac SMS idx=%s - ponowie probe", index)


def load_schedule():
    if not os.path.exists(XLSX_FILE):
        return []

    wb = load_workbook(XLSX_FILE, data_only=True)
    ws = wb.active
    schedule = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 6:
            continue

        sd, ed, st, et, _, num = row[:6]
        if not all([sd, ed, st, et]):
            continue

        number = str(num).strip()
        if not PHONE_REGEX.fullmatch(number):
            continue

        start = LOCAL_TZ.localize(datetime.datetime.combine(sd, st))
        end = LOCAL_TZ.localize(datetime.datetime.combine(ed, et))

        schedule.append((start, end, number))

    return schedule


def find_active_forward(schedule):
    now = datetime.datetime.now(LOCAL_TZ)
    for start, end, number in schedule:
        if start <= now <= end:
            return number
    return None


def main():
    ser = None
    last_number = None
    last_schedule_check = 0

    while True:

        if ser is None or not ser.is_open:
            ser = open_modem()
            last_number = None

            if ser:
                send_at(ser, "AT")
                send_at(ser, "ATE0")
                send_at(ser, "AT+CLIP=1")
                send_at(ser, "AT+CMGF=1")
                send_at(ser, 'AT+CSCS="GSM"', log_cmd=True)
                send_at(ser, 'AT+CPMS="ME","ME","ME"', log_cmd=True)
            else:
                time.sleep(5)
                continue

        try:
            if time.time() - last_schedule_check > SCHEDULE_CHECK_INTERVAL:
                schedule = load_schedule()
                number = find_active_forward(schedule)

                if number and number != last_number:
                    enable_forwarding(ser, number)
                    last_number = number
                elif not number and last_number:
                    disable_forwarding(ser)
                    last_number = None

                if last_number:
                    log.info("Aktualne przekierowanie: %s", last_number)
                else:
                    log.info("Brak aktywnego przekierowania")

                last_schedule_check = time.time()

            if last_number:
                process_all_sms(ser, last_number)

        except Exception as e:
            log.error("Blad petli glownej: %s", e)
            try:
                ser.close()
            except:
                pass
            ser = None

        time.sleep(SMS_CHECK_INTERVAL)


if __name__ == "__main__":
    main()
