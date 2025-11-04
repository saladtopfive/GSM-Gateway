import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

FILE_PATH = "schedule_data.xlsx"

# Tworzy nowy plik Excela, jeśli nie istnieje
def create_excel_file():
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Harmonogram"
        ws.append(["Data od", "Data do", "Godzina od", "Godzina do", "Numer docelowy", "Aktywne"])
        wb.save(FILE_PATH)

def load_schedules():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    wb.close()
    return data

def save_schedules(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Harmonogram"
    ws.append(["Data od", "Data do", "Godzina od", "Godzina do", "Numer docelowy", "Aktywne"])
    for row in data:
        ws.append(row)
    wb.save(FILE_PATH)

def add_schedule():
    d_od = entry_date_from.get()
    d_do = entry_date_to.get()
    g_od = entry_time_from.get()
    g_do = entry_time_to.get()
    numer = entry_number.get()
    aktywne = var_active.get()

    if not all([d_od, d_do, g_od, g_do, numer]):
        messagebox.showwarning("Błąd", "Wypełnij wszystkie pola.")
        return

    try:
        datetime.strptime(d_od, "%Y-%m-%d")
        datetime.strptime(d_do, "%Y-%m-%d")
    except ValueError:
        messagebox.showwarning("Błąd", "Nieprawidłowy format daty. Użyj RRRR-MM-DD.")
        return

    data = load_schedules()
    data.append((d_od, d_do, g_od, g_do, numer, "TAK" if aktywne else "NIE"))
    save_schedules(data)
    refresh_table()
    clear_inputs()

def delete_schedule():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Błąd", "Wybierz wpis do usunięcia.")
        return
    index = tree.index(selected[0])
    data = load_schedules()
    data.pop(index)
    save_schedules(data)
    refresh_table()

def refresh_table():
    for row in tree.get_children():
        tree.delete(row)
    for row in load_schedules():
        tree.insert("", "end", values=row)

def clear_inputs():
    entry_date_from.delete(0, tk.END)
    entry_date_to.delete(0, tk.END)
    entry_time_from.delete(0, tk.END)
    entry_time_to.delete(0, tk.END)
    entry_number.delete(0, tk.END)
    var_active.set(False)

# --- GUI ---
root = tk.Tk()
root.title("🗓 Edytor harmonogramu przekierowań")
root.geometry("1100x600")  # większe okno
root.option_add("*Font", "Arial 12")

frame_inputs = ttk.Frame(root, padding=15)
frame_inputs.pack(fill="x")

# Pola wejściowe
ttk.Label(frame_inputs, text="Data od (YYYY-MM-DD):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
entry_date_from = ttk.Entry(frame_inputs, width=15, font=("Arial", 12))
entry_date_from.grid(row=0, column=1, padx=5, pady=5)

ttk.Label(frame_inputs, text="Data do:").grid(row=0, column=2, sticky="e", padx=5, pady=5)
entry_date_to = ttk.Entry(frame_inputs, width=15, font=("Arial", 12))
entry_date_to.grid(row=0, column=3, padx=5, pady=5)

ttk.Label(frame_inputs, text="Godzina od (HH:MM):").grid(row=1, column=0, sticky="e", padx=5, pady=5)
entry_time_from = ttk.Entry(frame_inputs, width=15, font=("Arial", 12))
entry_time_from.grid(row=1, column=1, padx=5, pady=5)

ttk.Label(frame_inputs, text="Godzina do:").grid(row=1, column=2, sticky="e", padx=5, pady=5)
entry_time_to = ttk.Entry(frame_inputs, width=15, font=("Arial", 12))
entry_time_to.grid(row=1, column=3, padx=5, pady=5)

ttk.Label(frame_inputs, text="Numer docelowy:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
entry_number = ttk.Entry(frame_inputs, width=20, font=("Arial", 12))
entry_number.grid(row=2, column=1, padx=5, pady=5)

var_active = tk.BooleanVar()
ttk.Checkbutton(frame_inputs, text="Aktywne", variable=var_active).grid(row=2, column=2, padx=5, pady=5)

# Przyciski
ttk.Button(frame_inputs, text="➕ Dodaj", command=add_schedule).grid(row=3, column=0, pady=10)
ttk.Button(frame_inputs, text="❌ Usuń", command=delete_schedule).grid(row=3, column=1, pady=10)
ttk.Button(frame_inputs, text="🔄 Odśwież", command=refresh_table).grid(row=3, column=2, pady=10)

# Tabela
frame_table = ttk.Frame(root, padding=10)
frame_table.pack(fill="both", expand=True)

columns = ("Data od", "Data do", "Godzina od", "Godzina do", "Numer docelowy", "Aktywne")
tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=15)
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=150, anchor="center")

# Scrollbar
scrollbar_y = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
tree.configure(yscroll=scrollbar_y.set)
scrollbar_y.pack(side="right", fill="y")
tree.pack(fill="both", expand=True)

create_excel_file()
refresh_table()

root.mainloop()
