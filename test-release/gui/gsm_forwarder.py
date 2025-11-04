#!/usr/bin/env python3
import serial
import time
from datetime import datetime
from openpyxl import load_workbook
import os

# === KONFIGURACJA MODEMU ===
PORT = "/dev/ttyUSB2"
BAUD = 115200
SLEEP_INTERVAL = 60  # co ile sekund sprawdzać harmonogram

# === ŚCIEŻKA DO HARMONOGRAMU ===
FILE_PATH = "schedule_data.xlsx"

# === FUNKCJE ===
def send_command(ser, cmd, delay=1):
    """Wysyła komendę AT i zwraca odpowiedź modemu."""
    ser.write((cmd + "\r").encode())
    time.sleep(delay)
    resp = ser.read_all().decode(errors="ignore").strip()
    print(f"> {cmd}\n{resp}\n")
    return resp

def set_forward(ser, number):
    """Ustawia przekierowanie bezwarunkowe (voice)."""
    print(f"📞 Ustawiam przekierowanie na {number}...")
    return send_command(ser, f'AT+CCFC=0,3,"{number}",145')

def disable_forward(ser):
    """Wyłącza przekierowanie."""
    print("⛔ Wyłączam przekierowanie...")
    return send_command(ser, 'AT+CCFC=0,0')

def get_current_target():
    """Sprawdza harmonogram i zwraca numer, który jest aktywny w tej chwili."""
    if not os.path.exists(FILE_PATH):
        print("⚠️  Brak pliku harmonogramu, brak przekierowania.")
        return None

    try:
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        now = datetime.now()
        for row in ws.iter_rows(min_row=2, values_only=True):
            d_od, d_do, g_od, g_do, numer, aktywne = row
            if not (d_od and d_do and g_od and g_do and numer):
                continue
            if str(aktywne).upper() not in ["TAK", "TRUE", "1", "YES"]:
                continue

            try:
                date_from = datetime.strptime(str(d_od), "%Y-%m-%d")
                date_to = datetime.strptime(str(d_do), "%Y-%m-%d")
                time_from = datetime.strptime(str(g_od), "%H:%M").time()
                time_to = datetime.strptime(str(g_do), "%H:%M").time()
            except ValueError:
                continue

            if date_from.date() <= now.date() <= date_to.date():
                if time_from <= now.time() <= time_to:
                    return numer
        wb.close()
    except Exception as e:
        print(f"❌ Błąd podczas odczytu harmonogramu: {e}")
        return None
    return None

# === GŁÓWNY PROGRAM ===
def main():
    print("🔌 Łączenie z modemem...")
    ser = serial.Serial(PORT, BAUD, timeout=2)

    send_command(ser, "AT", 0.5)
    send_command(ser, "ATE0", 0.2)
    send_command(ser, "AT+CLIP=1", 0.2)

    current_number = None

    try:
        while True:
            new_number = get_current_target()
            if new_number != current_number:
                if new_number:
                    set_forward(ser, new_number)
                else:
                    disable_forward(ser)
                current_number = new_number
            print(f"🕓 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Aktualne przekierowanie: {current_number or 'Brak'}")
            time.sleep(SLEEP_INTERVAL)

    except KeyboardInterrupt:
        print("⛔ Zakończenie programu, wyłączam przekierowanie...")
        disable_forward(ser)
        ser.close()
        print("✅ Zakończono.")

if __name__ == "__main__":
    main()
