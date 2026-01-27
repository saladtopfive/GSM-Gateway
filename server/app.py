from flask import Flask, request, render_template, jsonify, send_file
from openpyxl import load_workbook
from datetime import datetime
import pytz
import os
import shutil
import tempfile
import re

app = Flask(__name__)

LOCAL_TZ = pytz.timezone("Europe/Warsaw")

UPLOAD_PATH = "/home/kp_rpi_user/GSM-Gateway/stable-release/schedule.xlsx"
ALLOWED_EXTENSIONS = {"xlsx"}

PHONE_REGEX = re.compile(r"^(?:\+48)?\d{9}$")

# ===================== HELPERS =====================

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def load_schedule():
    if not os.path.exists(UPLOAD_PATH):
        return []

    wb = load_workbook(UPLOAD_PATH, data_only=True)
    ws = wb.active
    rows = []

    for row_idx in range(2, ws.max_row + 1):
        values = [ws[f"{c}{row_idx}"].value for c in ("A", "B", "C", "D", "E", "F")]

        # twardy STOP
        if all(v is None for v in values):
            break

        if any(v is None for v in values):
            continue

        sd, ed, st, et, name, number = values

        try:
            start = LOCAL_TZ.localize(datetime.combine(sd, st))
            end = LOCAL_TZ.localize(datetime.combine(ed, et))
        except Exception:
            continue

        number = str(number).strip()
        if number.startswith(("'", "’", "‘")):
            number = number[1:]

        if not PHONE_REGEX.match(number):
            continue

        rows.append((start, end, name, number))

    return sorted(rows, key=lambda x: x[0])

# ===================== ROUTES =====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/status")
def status():
    now = datetime.now(LOCAL_TZ)
    schedule = load_schedule()

    current = None
    next_one = None

    for start, end, name, number in schedule:
        if start <= now <= end:
            current = (name, number, start, end)
        elif start > now and next_one is None:
            next_one = (name, number, start, end)

    def fmt(entry):
        if not entry:
            return None

        name, number, start, end = entry
        return {
            "person": name,
            "number": number,
            "start": start.strftime("%Y-%m-%d %H:%M"),
            "end": end.strftime("%Y-%m-%d %H:%M"),
        }

    return jsonify({
        "current": fmt(current),
        "next": fmt(next_one)
    })


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "Brak pliku w żądaniu"}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "Nie wybrano pliku"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Dozwolony jest tylko plik .xlsx"}), 400

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        expected = ["start_date", "end_date", "start_time", "end_time", "name", "number"]

        if headers[:6] != expected:
            os.unlink(tmp_path)
            return jsonify({"error": "Nieprawidłowa struktura pliku Excel"}), 400

    except Exception:
        os.unlink(tmp_path)
        return jsonify({"error": "Nie udało się odczytać pliku Excel"}), 400

    shutil.move(tmp_path, UPLOAD_PATH)
    return jsonify({"status": "ok"})


@app.route("/download")
def download():
    if not os.path.exists(UPLOAD_PATH):
        return jsonify({"error": "Brak pliku harmonogramu"}), 404

    return send_file(
        UPLOAD_PATH,
        as_attachment=True,
        download_name="schedule.xlsx"
    )

# ===================== ENTRY =====================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
